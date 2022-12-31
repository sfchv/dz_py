import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pdfkit
import datetime as dt
import os


class UserInput:
    file_name: str
    profession_name: str
    area_name: str

    def __init__(self, file_name: str = None, profession_name: str = None, area_name: str = None):
        if file_name is not None:
            self.file_name = file_name
        else:
            self.file_name = self._get_correct_input("Введите название файла: ", 'str')

        if profession_name is not None:
            self.profession_name = profession_name
        else:
            self.profession_name = self._get_correct_input(
                "Введите название профессии (можно ввести несколько через разделитель '|'): ", 'str')

        if area_name is not None:
            self.area_name = area_name
        else:
            self.area_name = self._get_correct_input("Введите название региона: ", 'str')

    def _get_correct_input(self, question: str, input_type: str,
                           error_msg: str = "Данные некорректны, повторите ввод.") -> str:
        user_input = input(question)
        while not self._is_correct_input(user_input, input_type):
            print(error_msg)
            user_input = input(question)
        return user_input

    @staticmethod
    def _is_correct_input(user_input: str, input_type: str) -> bool:
        if user_input == "" or input_type == "":
            return False
        elif input_type == 'bool':
            return user_input.lower() in ['да', 'нет']
        elif input_type == 'int':
            return user_input.isdigit()
        return True


class Report:
    """Класс формирования отчёта по данным из DataSet.

    Attributes
    ----------
    workbook : Workbook
        Класс, содержащий в себе функционал для работы с Excel таблицей.
    data : dict
        Словарь данных, получаемый из DataSet.
    """
    workbook: Workbook
    data: dict
    profession_name: str

    def __init__(self, data: dict, profession_name: str):
        """Инициализирует объект Report. Создаёт пустой Workbook, распаковывает kwargs.

        :param data: Словарь с данными из DataSet.
        """
        self.workbook = Workbook()
        self.data = data
        professions = profession_name.split('|')
        if len(professions) > 1:
            self.profession_name = f'\"{professions[0]}\" и ещё {len(professions) - 1}'
        else:
            self.profession_name = f'\"{profession_name}\"'

    # region Excel
    def generate_excel(self, file_name: str) -> None:
        """
        Генерирует и сохраняет Excel-файл.

        :param file_name: название Excel-файла с явно указанным расширением.
        """
        self.fill_with_statistics()
        self.workbook.save(file_name)

    def fill_with_statistics(self) -> None:
        """
        Заполняет два листа Excel-файла статистикой.

        """
        self.fill_salaries_statistics()
        self.fill_cities_statistics()

    def fill_salaries_statistics(self) -> None:
        """
        Заполняет первую страницу данными о годах, зарплатах и количествах вакансий.

        """
        ws = self.workbook.active
        ws.title = 'Статистика по годам'
        salaries_by_years = self.data["Уровень зарплат по годам"][0]
        vacancies_by_years = self.data["Количество вакансий по годам"][0]
        profession_salaries_by_years = self.data["Уровень зарплат по годам"][1]
        profession_vacancies_by_years = self.data["Количество вакансий по годам"][1]

        self.fill_column('Год', list(salaries_by_years.keys()),
                         [cell[0] for cell in ws['A1':f'A{len(salaries_by_years) + 1}']])

        self.fill_column('Средняя зарплата', list(salaries_by_years.values()),
                         [cell[0] for cell in ws['B1':f'B{len(salaries_by_years) + 1}']])
        self.fill_column(f'Средняя зарплата - {self.profession_name}',
                         list(profession_salaries_by_years.values()),
                         [cell[0] for cell in ws['C1':f'C{len(profession_salaries_by_years) + 1}']])

        self.fill_column('Количество вакансий', list(vacancies_by_years.values()),
                         [cell[0] for cell in ws['D1':f'D{len(vacancies_by_years) + 1}']])
        self.fill_column(f'Количество вакансий - {self.profession_name}',
                         list(profession_vacancies_by_years.values()),
                         [cell[0] for cell in ws['E1':f'E{len(profession_vacancies_by_years) + 1}']])

        self.update_worksheet_settings(ws)

    def fill_cities_statistics(self) -> None:
        """
        Создаёт и переключается на второй лист Excel-файла. Заполняет его данными о городах и зарплатах.

        """
        self.workbook.create_sheet("Статистика по городам")
        ws = self.workbook["Статистика по городам"]
        salaries_by_cities = self.data["Уровень зарплат по городам"]
        vacs_ratio_by_cities = self.data["Доля вакансий по городам"]

        self.fill_column('Город', list(salaries_by_cities.keys()),
                         [cell[0] for cell in ws['A1':f'A{len(salaries_by_cities) + 1}']])
        self.fill_column('Уровень зарплат', list(salaries_by_cities.values()),
                         [cell[0] for cell in ws['B1': f'B{len(salaries_by_cities) + 1}']])

        self.fill_column('Город', list(vacs_ratio_by_cities.keys()),
                         [cell[0] for cell in ws['D1':f'D{len(vacs_ratio_by_cities) + 1}']])
        self.fill_column('Доля вакансий', list(vacs_ratio_by_cities.values()),
                         [cell[0] for cell in ws['E1': f'E{len(vacs_ratio_by_cities) + 1}']])

        self.set_column_percent([cell[0] for cell in ws['E2': f'E{len(vacs_ratio_by_cities) + 1}']])
        self.update_worksheet_settings(ws)

    @staticmethod
    def fill_column(header: str, data: list, column_cells: list) -> None:
        """
        Заполняет столбец данными.

        :param header: Заголовок столбца, записывается в первой ячейке из списка клеток.
        :param data: Данные для записи в клетки.
        :param column_cells: Список клеток, в которые будут записаны данные.
        """
        column_cells[0].value = header
        for cell, value in zip(column_cells[1:], data):
            cell.value = value

    @staticmethod
    def set_column_percent(column: list) -> None:
        """
        Устанавливает процентный формат для всех ячеек в этом столбце.
        """
        for cell in column:
            cell.number_format = FORMAT_PERCENTAGE_00

    def update_worksheet_settings(self, ws) -> None:
        """
        Устанавливает границы и ширины для страницы Excel-файла.

        :param ws: страница Excel-файла.
        """
        self.set_borders(ws)
        self.set_column_width(ws)

    @staticmethod
    def set_borders(ws) -> None:
        """
        Устанавливает границы для всех непустых клеток и жирный шрифт для заголовков столбцов.

        :param ws: страница Excel-файла.
        """
        is_first_row = True
        for row in ws.rows:
            for cell in row:
                if not cell.value:
                    continue
                cell.border = Border(top=Side(border_style="thin", color="000000"),
                                     left=Side(border_style="thin", color="000000"),
                                     right=Side(border_style="thin", color="000000"),
                                     bottom=Side(border_style="thin", color="000000"))
                if is_first_row:
                    cell.font = Font(bold=True)
            is_first_row = False

    @staticmethod
    def set_column_width(ws) -> None:
        """
        Устанавливает ширину столбца, ориентируясь на максимально большую ячейку в нём.

        :param ws: страница Excel-файла.
        """
        a = {0: "A", 1: "B", 2: "C", 3: "D", 4: "E", 6: "F", 7: "G"}
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value)) + 1))

        for col, value in dims.items():
            ws.column_dimensions[a[col - 1]].width = value

    # endregion
    # region Plot

    def generate_image(self, file_name: str, show_result: bool = False) -> None:
        """
        Генерирует и сохраняет изображение с графиками, на основе данных data.

        :param file_name: Название для сохранения изображения.
        :param show_result: Показывать ли изображение после генерации. По-умолчанию False.
        """
        self.draw_graphs()
        plt.tight_layout()
        plt.savefig(file_name, dpi=300)
        if show_result:
            plt.show()

    def draw_graphs(self) -> None:
        """
        Рисует 4 графика на сетке 2x2. Каждый график строится на основании данных каждого ключа из data.

        """
        # figure, (ax1, ax2) = plt.subplots(2)
        figure, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2)
        self.draw_bar_graph(ax1, "Уровень зарплат по годам")
        self.draw_bar_graph(ax2, "Количество вакансий по годам")
        self.draw_invert_bar_graph(ax3, "Уровень зарплат по городам")
        self.draw_pie_graph(ax4, "Доля вакансий по городам")

    def draw_bar_graph(self, subplot, name: str) -> None:
        """
        Рисует столбчатую диаграмму.

        :param subplot: Подобласть для отрисовки графика.
        :param name: Название графика. Должен соответствовать ключу из data.
        """
        bar_width = 0.4
        first_label = 'средняя з/п'
        second_label = f'з/п {self.profession_name}'
        if name == "Количество вакансий по годам":
            first_label = "Количество вакансий"
            second_label = f"Количество вакансий\n{self.profession_name}"

        average_by_years: dict = self.data[name][0]
        profession_average_by_years: dict = self.data[name][1]

        x_axis = np.arange(len(average_by_years.keys()))

        subplot.bar(x_axis - bar_width / 2, average_by_years.values(), width=bar_width, label=first_label)
        subplot.bar(x_axis + bar_width / 2, profession_average_by_years.values(),
                    width=bar_width, label=second_label)
        subplot.set_xticks(x_axis, average_by_years.keys())
        subplot.set_xticklabels(average_by_years.keys(), rotation='vertical', va='top', ha='center')

        subplot.set_title(name)
        subplot.grid(True, axis='y')
        subplot.tick_params(axis='both', labelsize=8)
        subplot.legend(fontsize=8)

    def draw_invert_bar_graph(self, subplot, name: str) -> None:
        """
        Рисует повёрнутую на левый бок столбчатую диаграмму.

        :param subplot: Подобласть для отрисовки графика.
        :param name: Название графика. Должен соответствовать ключу из data.
        """
        subplot.invert_yaxis()
        courses = list(self.data[name].keys())
        courses = [label.replace(' ', '\n').replace('-', '-\n') for label in courses]
        values = list(self.data[name].values())
        subplot.barh(courses, values)
        subplot.set_yticklabels(courses, fontsize=6, va='center', ha='right')

        subplot.set_title(name)
        subplot.grid(True, axis='x')
        subplot.tick_params(axis='both', labelsize=8)

    def draw_pie_graph(self, subplot, name: str) -> None:
        """
        Рисует круговую диаграмму.

        :param subplot: Подобласть для отрисовки графика.
        :param name: Название графика. Должен соответствовать ключу из data.
        """
        data = self.data[name]
        other = 1 - sum((list(data.values())))
        new_dic = {'Другие': other}
        new_dic.update(data)

        labels = list(new_dic.keys())
        sizes = list(new_dic.values())

        subplot.set_title(name)
        subplot.pie(sizes, labels=labels, textprops={'fontsize': 6})
        subplot.axis('scaled')

    # endregion
    # region PDF

    def generate_pdf(self, name: str):
        """
        Генерирует PDF-файл на основании данных из DataSet - data и разметки PDF - pdf_template.html.

        :param name: Название сохраняемого PDF-файла с явно указанным расширением.
        """

        excel_file_name = "report.xlsx"
        image_file_name = "graph.png"
        self.generate_image(image_file_name)
        self.generate_excel(excel_file_name)

        image_file = os.path.join(os.path.dirname(name), image_file_name)
        header_year = ["Год", "Средняя зарплата", f"Средняя зарплата - {self.profession_name}",
                       "Количество вакансий",
                       f"Количество вакансий - {self.profession_name}"]
        header_city = ["Город", "Уровень зарплат", '', "Город", "Доля вакансий"]

        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")

        salaries_by_years = self.data["Уровень зарплат по годам"][0]
        vacancies_by_years = self.data["Количество вакансий по годам"][0]
        profession_salaries_by_years = self.data["Уровень зарплат по годам"][1]
        profession_vacancies_by_years = self.data["Количество вакансий по годам"][1]
        salaries_by_cities = self.data["Уровень зарплат по городам"]
        ratio_vacancy_by_cities = {city: str(f'{ratio * 100:,.2f}%').replace('.', ',')
                                   for city, ratio in self.data["Доля вакансий по городам"].items()}

        salary_data = {year: [salary, count, salary_vac, count_vac]
                       for year, salary, count, salary_vac, count_vac in zip(salaries_by_years.keys(),
                                                                             salaries_by_years.values(),
                                                                             vacancies_by_years.values(),
                                                                             profession_salaries_by_years.values(),
                                                                             profession_vacancies_by_years.values())}

        city_data = {index: [salary_city, salary, ratio_city, ratio]
                     for index, (salary_city, salary, ratio_city, ratio) in
                     enumerate(zip(salaries_by_cities.keys(),
                                   salaries_by_cities.values(),
                                   ratio_vacancy_by_cities.keys(),
                                   ratio_vacancy_by_cities.values()))}

        pdf_template = template.render(
            {'image_file': image_file,
             'image_style': 'style="max-width:1024px; max-height:680px"',
             'salary_data': salary_data,
             'city_data': city_data,
             'header_year': header_year,
             'header_city': header_city,
             'profession_name': f"{self.profession_name}",
             'h1_style': 'style="text-align:center; font-size:32px"',
             'h2_style': 'style="text-align:center"',
             'cell_style_none': "style=''",
             'cell_style': 'style="border:1px solid black; border-collapse: collapse; font-size: 16px; height: 19pt;'
                           'padding: 5px; text-align:center"'})

        config = pdfkit.configuration(wkhtmltopdf=r'D:\Programs\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, name, configuration=config, options={'enable-local-file-access': None})
    # endregion


def get_salaries_by_years(df: pd.DataFrame) -> dict:
    return (df
            .groupby("published_at")["salary"].agg(np.sum)
            .floordiv(df
                      .groupby('published_at')['published_at']
                      .count())
            .to_dict()
            )


def get_vacancies_count_by_years(df: pd.DataFrame) -> dict:
    return (
        df
        .value_counts("published_at", ascending=True)
        .sort_index()
        .to_dict()
    )


def check_substring_in_name(df: pd.DataFrame, name: str) -> pd.DataFrame:
    return df.loc[
        df["name"]
        .apply(lambda s: s.lower())
        .str.contains(name.lower())
    ]


def get_data_by_years(df: pd.DataFrame, profession_name: str) -> dict:
    salaries_by_years = df.pipe(get_salaries_by_years)
    vacancies_count_by_years = df.pipe(get_vacancies_count_by_years)
    profession_salaries_by_years = check_substring_in_name(df, profession_name).pipe(get_salaries_by_years)
    profession_vacancies_count_by_years = check_substring_in_name(df, profession_name).pipe(
        get_vacancies_count_by_years)
    return {
        "Уровень зарплат по годам": (salaries_by_years, profession_salaries_by_years),
        "Количество вакансий по годам": (vacancies_count_by_years, profession_vacancies_count_by_years)
    }


def get_salaries_by_cities(df: pd.DataFrame) -> dict:
    return (df
            .groupby("area_name")["salary"].agg(sum)
            .floordiv(df
                      .groupby('area_name')['area_name']
                      .count())
            .nlargest(10)
            .to_dict()
            )


def get_cities_more_percent_series(df: pd.DataFrame) -> pd.Series:
    return (
        df["area_name"]
        .value_counts(normalize=True)
        .loc[lambda x: x > 0.01]
    )


def get_cities_more_percent(df: pd.DataFrame) -> pd.DataFrame:
    return df.loc[
        df["area_name"].isin(
            df.pipe(get_cities_more_percent_series)
            .index)
    ]


def get_data_by_cities(df: pd.DataFrame) -> dict:
    salaries_by_cities = get_salaries_by_cities(df.pipe(get_cities_more_percent))
    ratio_vacancies_by_cities = (df
                                 .pipe(get_cities_more_percent_series)
                                 .nlargest(10)
                                 .to_dict()
                                 )

    return {
        "Уровень зарплат по городам": salaries_by_cities,
        "Доля вакансий по городам": ratio_vacancies_by_cities
    }


def main() -> None:
    hh_ru = "../API/hh_ru_joined_salary.csv"
    big = "vacancies_joined_salary.csv"
    ui = UserInput(big, "Программист|IT|Senior|Middle|Junior|Аналитик", "Москва")
    df = pd.read_csv(ui.file_name, dtype={"name": "str", "salary": "Int64", "area_name": "str"}, verbose=True)
    df = df.assign(published_at=df['published_at'].apply(lambda s: dt.datetime.fromisoformat(s).year).astype("int32"))

    data = get_data_by_years(df, ui.profession_name)
    data.update(df.pipe(get_data_by_cities))

    report = Report(data, ui.profession_name)
    report.generate_pdf("report.pdf")


if __name__ == "__main__":
    main()
